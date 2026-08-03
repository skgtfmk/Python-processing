
import re
import datetime
from openpyxl import load_workbook

import re
from pathlib import Path

def get_protocol_and_end_timestamp(rows):
    protocol_file = None
    end_date = None
    end_time = None

    for row in rows:
        for cell in row:
            if cell is None:
                continue

            value = str(cell).strip()

            if not protocol_file:
                match = re.search(r"Protocol File Name\s*=\s*(.+)", value)
                if match:
                    protocol_path = match.group(1).strip()
                    protocol_file = Path(protocol_path).stem

            if not end_date:
                match = re.search(r"End Date\s*=\s*(.+)", value)
                if match:
                    end_date = match.group(1).strip()

            if not end_time:
                match = re.search(r"End Time\s*=\s*(.+)", value)
                if match:
                    end_time = match.group(1).strip()

            if protocol_file and end_date and end_time:
                return protocol_file, f"{end_date} {end_time}"

    return protocol_file, None

def find_Plate_ID (fContents):
    current_r =0
    all_found = False
    lPlate_Id = ''
    lPlate_date = ''
    lplate_time = ''
    while current_r < len(fContents) :
        if lPlate_Id == '':
            re_prop = re.search('Protocol File Name = (.*).fmp',','.join(fContents[current_r]))
            if re_prop:
        #       logger.info(re_prop.group(1), re_prop.group(2))
                try:
                    lPlate_Id = re_prop.group(2)
                except:
                    lPlate_Id = 'Plate ' + str(plate_n)
        if lPlate_date == '':
            re_prop = re.search('End Date = (.*)',','.join(fContents[current_r]))
            if re_prop:
        #       logger.info(re_prop.group(1), re_prop.group(2))
                try:
                    lPlate_date = re_prop.group(2)
                except:
                    lPlate_date = datetime.datetime.today().strftime('%Y-%m-%d')
        if lplate_time == '':
            re_prop = re.search('End Time = (.*)',','.join(fContents[current_r]))
            if re_prop:
        #       logger.info(re_prop.group(1), re_prop.group(2))
                try:
                    lplate_time = re_prop.group(2)
                except:
                    lplate_time = datetime.datetime.today().strftime('%H:%M:%S')
    return lPlate_Id, lPlate_date + ' ' + lplate_time

fPath = 'C:\\Users\\skgtfmk\OneDrive - University College London\\Documents\\Requests information\\B-Score plate correction\\FLIPR'
fFile = '12032025_03 dec 2025 -EAAT2 -CNS library  AM plate 1_n001.statAll.xlsx'
#df = openpyxl.load_workbook(fPath + '\\' + fFile)
#max_row = df['12032025_03 dec 2025 -EAAT2 -CN'].max_row
wb = load_workbook(fPath + '\\' + fFile, data_only=True)
ws = wb[wb.sheetnames[0]]

rows = list(ws.iter_rows(values_only=True))
#for r in rows:
#    print (r)

plate_id, plate_date = find_Plate_ID(rows)

#plate_id, plate_date = get_protocol_and_end_timestamp(fPath + '\\' + fFile)
#print(plate_id, plate_date)